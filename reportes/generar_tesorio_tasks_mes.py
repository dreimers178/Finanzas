#!/usr/bin/env python3
"""Reporte de Tesorio Tasks ENFOCADO en un mes.
Tabla general = últimos 3 meses; distribución/prioridad/lentas = mes destacado.

Uso: python3 generar_tesorio_tasks_mes.py <export.csv> <mes AAAA-MM> [salida.xlsx] [hoy AAAA-MM-DD]
"""
import sys
import pandas as pd
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

SRC=sys.argv[1]
TARGET=sys.argv[2] if len(sys.argv)>2 else None
OUT=sys.argv[3] if len(sys.argv)>3 else "/home/user/Finanzas/Reporte_Tesorio_Agosto_2026.xlsx"
TODAY=pd.Timestamp(sys.argv[4], tz='UTC') if len(sys.argv)>4 else pd.Timestamp('2026-09-02', tz='UTC')

df=pd.read_csv(SRC, dtype=str)
df['c']=pd.to_datetime(df['Created At'],utc=True,errors='coerce')
df['d']=pd.to_datetime(df['Completed At'],utc=True,errors='coerce')
df=df[df['c'].notna()].reset_index(drop=True)
done=df[(df['Status']=='DONE') & df['d'].notna()].copy()
done['h']=(done['d']-done['c']).dt.total_seconds()/3600; done['dias']=done['h']/24
done['mes']=done['c'].dt.strftime('%Y-%m')
if TARGET is None: TARGET=sorted(done['mes'].unique())[-1]
last3=sorted(done['mes'].unique())[-3:]
tgt=done[done['mes']==TARGET]

NAVY="1F3864"
hfill=PatternFill("solid",fgColor=NAVY); sfill=PatternFill("solid",fgColor="2E5496")
hfont=Font(name="Arial",bold=True,color="FFFFFF",size=11); sfont=hfont
cfont=Font(name="Arial",size=10); bfont=Font(name="Arial",bold=True,size=10)
kfont=Font(name="Arial",bold=True,size=14,color=NAVY)
center=Alignment(horizontal="center",vertical="center"); left=Alignment(horizontal="left",vertical="center")
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
red=PatternFill("solid",fgColor="F4CCCC"); yel=PatternFill("solid",fgColor="FFF2CC")

wb=Workbook()

# Datos (calculo) — solo tasks de los últimos 3 meses + abiertas, para que las fórmulas cuadren con el enfoque
keep=df[(df['c'].dt.strftime('%Y-%m').isin(last3)) | (df['Status'].isin(['TO_DO','WORKING']))].copy()
dz=wb.active; dz.title="Datos (calculo)"
dcols=["Invoice Number","Status","Priority","Assigned by","Created At","Completed At",
       "Resolution (hrs)","Resolution (dias)","Mes (creado)","Días abiertos"]
for c,h in enumerate(dcols,1):
    cell=dz.cell(1,c,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
r=2
for _,row in keep.iterrows():
    dz.cell(r,1,row['Invoice Number'] if pd.notna(row['Invoice Number']) else "")
    dz.cell(r,2,row['Status']); dz.cell(r,3,row['Priority'] if pd.notna(row['Priority']) else "")
    dz.cell(r,4,row['Assigned by'] if pd.notna(row['Assigned by']) else "")
    ts=dz.cell(r,5,row['c'].tz_localize(None).to_pydatetime()); ts.number_format='yyyy-mm-dd hh:mm'
    if pd.notna(row['d']):
        me=dz.cell(r,6,row['d'].tz_localize(None).to_pydatetime()); me.number_format='yyyy-mm-dd hh:mm'
    if row['Status']=='DONE' and pd.notna(row['d']):
        h=(row['d']-row['c']).total_seconds()/3600; dz.cell(r,7,round(h,1)); dz.cell(r,8,round(h/24,1))
    dz.cell(r,7).number_format='0.0'; dz.cell(r,8).number_format='0.0'
    dz.cell(r,9,row['c'].strftime('%Y-%m'))
    if row['Status'] in ('TO_DO','WORKING'):
        dz.cell(r,10, max((TODAY.date()-row['c'].date()).days,0)); dz.cell(r,10).number_format='0'
    for c in range(1,11):
        cc=dz.cell(r,c); cc.font=cfont; cc.border=border; cc.alignment=left if c in (1,2,3,4) else center
    r+=1
LAST=r-1
for i,w in enumerate([16,10,15,16,18,18,15,15,13,13],1): dz.column_dimensions[get_column_letter(i)].width=w
dz.freeze_panes="A2"
D="'Datos (calculo)'"; ST=f"{D}!$B$2:$B${LAST}"; PRI=f"{D}!$C$2:$C${LAST}"
HRS=f"{D}!$G$2:$G${LAST}"; DAYS=f"{D}!$H$2:$H${LAST}"; MES=f"{D}!$I$2:$I${LAST}"

ws=wb.create_sheet(f"Reporte Tesorio {TARGET}", 0); ws.sheet_view.showGridLines=False
def title(cell,txt): ws[cell]=txt; ws[cell].font=sfont; ws[cell].fill=sfill; ws[cell].alignment=left
def hdr(row,cols,sc=1):
    for i,h in enumerate(cols):
        cell=ws.cell(row,sc+i,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
def box(row,col,val,fmt='General',font=cfont,al=center):
    cell=ws.cell(row,col,val); cell.font=font; cell.alignment=al; cell.number_format=fmt; cell.border=border; return cell

ws['A1']=f"REPORTE DE TESORIO TASKS — {TARGET}"; ws['A1'].font=Font(name="Arial",bold=True,size=16,color=NAVY)
ws['A2']=f"Resolución = Completed At − Created At (DONE). Tabla mensual = últimos 3 meses. Distribución/prioridad/lentas filtradas a {TARGET}. Abiertas = snapshot al {TODAY.date()}."
ws['A2'].font=Font(name="Arial",italic=True,size=9,color="808080")

# KPIs del mes
title('A4',f"  RESUMEN {TARGET}")
med_t=round(tgt['dias'].median(),1)
kpis=[("Creadas en el mes",f'=COUNTIF({MES},"{TARGET}")','0'),
 ("Completadas (DONE)",f'=COUNTIFS({MES},"{TARGET}",{HRS},">=0")','0'),
 ("Prom. resolución (días)",f'=IFERROR(ROUND(AVERAGEIF({MES},"{TARGET}",{DAYS}),1),0)','0.0'),
 ("Mediana (días)",med_t,'0.0'),
 ("% mismo día",f'=IFERROR(COUNTIFS({MES},"{TARGET}",{HRS},"<24")/COUNTIFS({MES},"{TARGET}",{HRS},">=0"),0)','0.0%'),
 ("Pendientes (TO_DO) hoy",f'=COUNTIF({ST},"TO_DO")','0'),
 ("En proceso (WORKING) hoy",f'=COUNTIF({ST},"WORKING")','0')]
r=5
for lbl,frm,fmt in kpis:
    ws.cell(r,1,lbl).font=cfont; ws.cell(r,1).alignment=left; ws.cell(r,1).border=border
    box(r,2,frm,fmt,kfont); r+=1
ws.column_dimensions['A'].width=27; ws.column_dimensions['B'].width=14

# Tabla 1: por mes (últimos 3)
title('D4',"  RESOLUTION TIME POR MES (últimos 3 meses)")
hdr(5,["Mes","# Creadas","Completadas","Prom. horas","Prom. días","Mediana días","% mismo día"],4)
med_mes=done.groupby('mes')['dias'].median().round(2).to_dict()
r=6
for m in last3:
    box(r,4,m,'General',bfont,left)
    box(r,5,f'=COUNTIF({MES},"{m}")','0')
    box(r,6,f'=COUNTIFS({MES},"{m}",{HRS},">=0")','0')
    box(r,7,f'=IFERROR(ROUND(AVERAGEIF({MES},"{m}",{HRS}),1),0)','0.0')
    box(r,8,f'=IFERROR(ROUND(AVERAGEIF({MES},"{m}",{DAYS}),1),0)','0.0')
    box(r,9, med_mes.get(m,0),'0.0')
    box(r,10,f'=IFERROR(COUNTIFS({MES},"{m}",{HRS},"<24")/COUNTIFS({MES},"{m}",{HRS},">=0"),0)','0.0%')
    r+=1
d3=done[done['mes'].isin(last3)]
box(r,4,"TOTAL",'General',bfont,left); box(r,5,f'=SUM(E6:E{r-1})','0',bfont); box(r,6,f'=SUM(F6:F{r-1})','0',bfont)
box(r,7,f'=IFERROR(ROUND(AVERAGEIFS({HRS},{MES},">={last3[0]}"),1),0)','0.0',bfont)
box(r,8,f'=IFERROR(ROUND(AVERAGEIFS({DAYS},{MES},">={last3[0]}"),1),0)','0.0',bfont)
box(r,9,round(d3['dias'].median(),1),'0.0',bfont)
box(r,10,f'=IFERROR(COUNTIFS({MES},">={last3[0]}",{HRS},"<24")/COUNTIFS({MES},">={last3[0]}",{HRS},">=0"),0)','0.0%',bfont)
ws.cell(r+1,4,"Medianas sobre datos actuales.").font=Font(name="Arial",italic=True,size=8,color="808080")
for col,w in zip(['D','E','F','G','H','I','J'],[10,11,12,12,11,12,12]): ws.column_dimensions[col].width=w

# Tabla 2: distribución (mes)
title('A16',f"  DISTRIBUCIÓN POR TIEMPO DE RESOLUCIÓN — {TARGET}")
hdr(17,["Rango","# Tasks","% del total"])
buckets=[("Mismo día (< 24 h)",'"<24"'),("1 – 2 días",f'">=24",{HRS},"<48"'),("2 – 4 días",f'">=48",{HRS},"<96"'),
 ("4 – 7 días",f'">=96",{HRS},"<168"'),("Más de 7 días",'">=168"')]
r=18; first=r
for lbl,cond in buckets:
    box(r,1,lbl,'General',cfont,left); box(r,2,f'=COUNTIFS({MES},"{TARGET}",{HRS},{cond})','0')
    box(r,3,f'=IFERROR(B{r}/COUNTIFS({MES},"{TARGET}",{HRS},">=0"),0)','0.0%'); r+=1
box(r,1,"TOTAL",'General',bfont,left); box(r,2,f'=SUM(B{first}:B{r-1})','0',bfont)
box(r,3,f'=IFERROR(B{r}/COUNTIFS({MES},"{TARGET}",{HRS},">=0"),0)','0.0%',bfont)

# Tabla 3: prioridad (mes)
title('D16',f"  RESOLUTION TIME POR PRIORIDAD — {TARGET}")
hdr(17,["Prioridad","# Tasks","% del total","Prom. días","% mismo día"],4)
prios=["URGENT","HIGH_PRIORITY","NORMAL"]
TOT=f'COUNTIFS({MES},"{TARGET}",{HRS},">=0")'; r=18
for p in prios:
    box(r,4,p,'General',cfont,left)
    box(r,5,f'=COUNTIFS({PRI},"{p}",{MES},"{TARGET}",{HRS},">=0")','0')
    box(r,6,f'=IFERROR(COUNTIFS({PRI},"{p}",{MES},"{TARGET}",{HRS},">=0")/{TOT},0)','0.0%')
    box(r,7,f'=IFERROR(ROUND(AVERAGEIFS({DAYS},{PRI},"{p}",{MES},"{TARGET}"),1),0)','0.0')
    box(r,8,f'=IFERROR(COUNTIFS({PRI},"{p}",{MES},"{TARGET}",{HRS},"<24")/COUNTIFS({PRI},"{p}",{MES},"{TARGET}",{HRS},">=0"),0)','0.0%')
    r+=1
box(r,4,"TOTAL",'General',bfont,left); box(r,5,f'={TOT}','0',bfont); box(r,6,'100%','0%',bfont)
box(r,7,f'=IFERROR(ROUND(AVERAGEIF({MES},"{TARGET}",{DAYS}),1),0)','0.0',bfont)
box(r,8,f'=IFERROR(COUNTIFS({MES},"{TARGET}",{HRS},"<24")/{TOT},0)','0.0%',bfont)
ws.freeze_panes="A3"

# Pendientes y En Proceso (snapshot)
op=df[df['Status'].isin(['TO_DO','WORKING'])].copy()
op['ab']=op['c'].apply(lambda x:max((TODAY.date()-x.date()).days,0)); op=op.sort_values('ab',ascending=False)
pz=wb.create_sheet("Pendientes y En Proceso"); pz.sheet_view.showGridLines=False
pz['A1']="TASKS ABIERTAS — Pendientes (TO_DO) y En Proceso (WORKING)"; pz['A1'].font=Font(name="Arial",bold=True,size=14,color=NAVY)
pz['A2']=f"{len(op)} tasks sin completar al {TODAY.date()} (snapshot actual, todas las fechas)."; pz['A2'].font=Font(name="Arial",italic=True,size=9,color="808080")
cols=["Invoice Number","Status","Priority","Assigned by","Created At","Días abiertos"]
for c,h in enumerate(cols,1):
    cell=pz.cell(4,c,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
r=5
for _,row in op.iterrows():
    pz.cell(r,1,row['Invoice Number'] if pd.notna(row['Invoice Number']) else "")
    pz.cell(r,2,row['Status']); pz.cell(r,3,row['Priority'] if pd.notna(row['Priority']) else "")
    pz.cell(r,4,row['Assigned by'] if pd.notna(row['Assigned by']) else "")
    t=pz.cell(r,5,row['c'].tz_localize(None).to_pydatetime()); t.number_format='yyyy-mm-dd hh:mm'
    dd=pz.cell(r,6,int(row['ab'])); dd.font=bfont
    if row['ab']>=7: dd.fill=red
    elif row['ab']>=2: dd.fill=yel
    for c in range(1,7):
        cc=pz.cell(r,c); cc.border=border
        if c!=6: cc.font=cfont
        cc.alignment=left if c in (1,3,4) else center
    r+=1
for i,w in enumerate([16,12,15,16,18,14],1): pz.column_dimensions[get_column_letter(i)].width=w
pz.freeze_panes="A5"

# Tasks lentas del mes (>7d)
slow=tgt[tgt['dias']>=7].sort_values('dias',ascending=False)
sz=wb.create_sheet(f"Tasks lentas {TARGET} (>7d)"); sz.sheet_view.showGridLines=False
sz['A1']=f"TASKS DE {TARGET} QUE TARDARON MÁS DE 7 DÍAS"; sz['A1'].font=Font(name="Arial",bold=True,size=14,color=NAVY)
sz['A2']=f"{len(slow)} de {len(tgt)} tasks completadas de {TARGET} superaron los 7 días."; sz['A2'].font=Font(name="Arial",italic=True,size=9,color="808080")
cols=["Invoice Number","Priority","Assigned by","Created At","Completed At","Días de resolución"]
for c,h in enumerate(cols,1):
    cell=sz.cell(4,c,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
r=5
for _,row in slow.iterrows():
    sz.cell(r,1,row['Invoice Number'] if pd.notna(row['Invoice Number']) else "")
    sz.cell(r,2,row['Priority'] if pd.notna(row['Priority']) else ""); sz.cell(r,3,row['Assigned by'] if pd.notna(row['Assigned by']) else "")
    t=sz.cell(r,4,row['c'].tz_localize(None).to_pydatetime()); t.number_format='yyyy-mm-dd hh:mm'
    m=sz.cell(r,5,row['d'].tz_localize(None).to_pydatetime()); m.number_format='yyyy-mm-dd hh:mm'
    d=sz.cell(r,6); d.value=f'=(E{r}-D{r})'; d.number_format='0.0'; d.fill=red; d.font=bfont
    for c in range(1,7):
        cc=sz.cell(r,c); cc.border=border
        if c!=6: cc.font=cfont
        cc.alignment=left if c in (1,3) else center
    r+=1
if len(slow):
    sz.cell(r+1,5,"Promedio:").font=bfont
    av=sz.cell(r+1,6,f'=ROUND(AVERAGE(F5:F{r-1}),1)'); av.font=bfont; av.number_format='0.0'
for i,w in enumerate([16,15,16,18,18,16],1): sz.column_dimensions[get_column_letter(i)].width=w
sz.freeze_panes="A5"

wb.calculation=CalcProperties(calcId=124519, fullCalcOnLoad=True)
wb.save(OUT)
print("OK ->",OUT,"| mes:",TARGET,"| ult3:",last3,"| lentas:",len(slow),"| abiertas:",len(op))
