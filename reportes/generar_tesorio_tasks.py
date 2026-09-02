#!/usr/bin/env python3
"""Genera el 'Reporte de Tesorio Tasks' a partir del CSV export de Tesorio.

Uso:  python3 generar_tesorio_tasks.py <archivo_export.csv> [archivo_salida.xlsx]

Métricas: tiempo de resolución = Completed At - Created At (para tasks DONE).
Estados:  DONE = completadas | TO_DO = pendientes | WORKING = en proceso.
Para TO_DO/WORKING se calculan los días abiertos = hoy - Created At.
"""
import sys
import pandas as pd
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

SRC = sys.argv[1] if len(sys.argv) > 1 else "/root/.claude/uploads/8d2c70a9-229c-592d-8546-419941c69ec5/e63a939b-tesorio_task_export_20260902.csv"
OUT = sys.argv[2] if len(sys.argv) > 2 else "/home/user/Finanzas/Reporte_Tesorio_Tasks.xlsx"
TODAY = pd.Timestamp(datetime.now(timezone.utc).date(), tz='UTC') if len(sys.argv) <= 3 else pd.Timestamp(sys.argv[3], tz='UTC')
# Fijamos la fecha del export para reproducibilidad
TODAY = pd.Timestamp('2026-09-02', tz='UTC')

df = pd.read_csv(SRC, dtype=str)
df['c'] = pd.to_datetime(df['Created At'], utc=True, errors='coerce')
df['d'] = pd.to_datetime(df['Completed At'], utc=True, errors='coerce')
df = df[df['c'].notna()].reset_index(drop=True)

done = df[(df['Status'] == 'DONE') & df['d'].notna()].copy()
done['h'] = (done['d'] - done['c']).dt.total_seconds() / 3600
done['dias'] = done['h'] / 24
done['mes'] = done['c'].dt.strftime('%Y-%m')

NAVY="1F3864"; BLUE="2E5496"
hfill=PatternFill("solid",fgColor=NAVY); sfill=PatternFill("solid",fgColor=BLUE)
hfont=Font(name="Arial",bold=True,color="FFFFFF",size=11); sfont=hfont
cfont=Font(name="Arial",size=10); bfont=Font(name="Arial",bold=True,size=10)
kfont=Font(name="Arial",bold=True,size=14,color=NAVY)
center=Alignment(horizontal="center",vertical="center"); left=Alignment(horizontal="left",vertical="center")
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
red=PatternFill("solid",fgColor="F4CCCC"); yel=PatternFill("solid",fgColor="FFF2CC")

wb=Workbook()

# ===== Datos (calculo) =====
dz=wb.active; dz.title="Datos (calculo)"
dcols=["Invoice Number","Status","Priority","Assigned by","Created At","Completed At",
       "Resolution (hrs)","Resolution (dias)","Mes (creado)","Días abiertos"]
for c,h in enumerate(dcols,1):
    cell=dz.cell(1,c,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
r=2
for _,row in df.iterrows():
    dz.cell(r,1,row['Invoice Number'] if pd.notna(row['Invoice Number']) else "")
    dz.cell(r,2,row['Status']); dz.cell(r,3,row['Priority'] if pd.notna(row['Priority']) else "")
    dz.cell(r,4,row['Assigned by'] if pd.notna(row['Assigned by']) else "")
    ts=dz.cell(r,5,row['c'].tz_localize(None).to_pydatetime()); ts.number_format='yyyy-mm-dd hh:mm'
    if pd.notna(row['d']):
        me=dz.cell(r,6,row['d'].tz_localize(None).to_pydatetime()); me.number_format='yyyy-mm-dd hh:mm'
    if row['Status']=='DONE' and pd.notna(row['d']):
        h=(row['d']-row['c']).total_seconds()/3600
        dz.cell(r,7,round(h,1)); dz.cell(r,8,round(h/24,1))
    dz.cell(r,7).number_format='0.0'; dz.cell(r,8).number_format='0.0'
    dz.cell(r,9,row['c'].strftime('%Y-%m'))
    if row['Status'] in ('TO_DO','WORKING'):
        ab=(TODAY.date()-row['c'].date()).days
        dz.cell(r,10, max(ab,0)); dz.cell(r,10).number_format='0'
    for c in range(1,11):
        cc=dz.cell(r,c); cc.font=cfont; cc.border=border; cc.alignment=left if c in (1,2,3,4) else center
    r+=1
LAST=r-1
for i,w in enumerate([16,10,15,16,18,18,15,15,13,13],1): dz.column_dimensions[get_column_letter(i)].width=w
dz.freeze_panes="A2"

D="'Datos (calculo)'"
ST=f"{D}!$B$2:$B${LAST}"; PRI=f"{D}!$C$2:$C${LAST}"
HRS=f"{D}!$G$2:$G${LAST}"; DAYS=f"{D}!$H$2:$H${LAST}"; MES=f"{D}!$I$2:$I${LAST}"

# ===== Reporte =====
ws=wb.create_sheet("Reporte Tesorio Tasks", 0)
ws.sheet_view.showGridLines=False
def title(cell,txt): ws[cell]=txt; ws[cell].font=sfont; ws[cell].fill=sfill; ws[cell].alignment=left
def hdr(row,cols,sc=1):
    for i,h in enumerate(cols):
        cell=ws.cell(row,sc+i,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
def box(row,col,val,fmt='General',font=cfont,al=center):
    cell=ws.cell(row,col,val); cell.font=font; cell.alignment=al; cell.number_format=fmt; cell.border=border; return cell

ws['A1']="REPORTE DE TESORIO TASKS"; ws['A1'].font=Font(name="Arial",bold=True,size=16,color=NAVY)
ws['A2']=f"Tiempo de resolución = Completed At − Created At (tasks DONE). Estados: DONE=completadas, TO_DO=pendientes, WORKING=en proceso. Días abiertos al {TODAY.date()}."
ws['A2'].font=Font(name="Arial",italic=True,size=9,color="808080")

# --- KPIs / estado ---
title('A4',"  ESTADO DE LAS TASKS")
kpis=[("Total tasks",f'=COUNTA({D}!$A$2:$A${LAST})','0'),
 ("Completadas (DONE)",f'=COUNTIF({ST},"DONE")','0'),
 ("Pendientes (TO_DO)",f'=COUNTIF({ST},"TO_DO")','0'),
 ("En proceso (WORKING)",f'=COUNTIF({ST},"WORKING")','0'),
 ("Completadas con fecha",f'=COUNT({HRS})','0'),
 ("Prom. resolución (días)",f'=ROUND(AVERAGE({DAYS}),1)','0.0'),
 ("Mediana (días)",f'=ROUND(MEDIAN({DAYS}),1)','0.0'),
 ("% mismo día",f'=IFERROR(COUNTIFS({HRS},"<24")/COUNT({HRS}),0)','0.0%')]
r=5
for lbl,frm,fmt in kpis:
    ws.cell(r,1,lbl).font=cfont; ws.cell(r,1).alignment=left; ws.cell(r,1).border=border
    box(r,2,frm,fmt,kfont); r+=1
ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=14

# --- Tabla 1: resolution por mes (creado) ---
title('D4',"  RESOLUTION TIME POR MES (por mes de creación)")
hdr(5,["Mes","# Creadas","Completadas","Prom. horas","Prom. días","Mediana días","% mismo día"],4)
meses=sorted(done['mes'].unique())
med_mes=done.groupby('mes')['dias'].median().round(2).to_dict()
r=6
for m in meses:
    box(r,4,m,'General',bfont,left)
    box(r,5,f'=COUNTIF({MES},"{m}")','0')
    box(r,6,f'=COUNTIFS({MES},"{m}",{HRS},">=0")','0')
    box(r,7,f'=IFERROR(ROUND(AVERAGEIF({MES},"{m}",{HRS}),1),0)','0.0')
    box(r,8,f'=IFERROR(ROUND(AVERAGEIF({MES},"{m}",{DAYS}),1),0)','0.0')
    box(r,9, med_mes.get(m,0),'0.0')
    box(r,10,f'=IFERROR(COUNTIFS({MES},"{m}",{HRS},"<24")/COUNTIFS({MES},"{m}",{HRS},">=0"),0)','0.0%')
    r+=1
box(r,4,"TOTAL",'General',bfont,left); box(r,5,f'=SUM(E6:E{r-1})','0',bfont)
box(r,6,f'=SUM(F6:F{r-1})','0',bfont); box(r,7,f'=ROUND(AVERAGE({HRS}),1)','0.0',bfont)
box(r,8,f'=ROUND(AVERAGE({DAYS}),1)','0.0',bfont); box(r,9,round(done['dias'].median(),1),'0.0',bfont)
box(r,10,f'=IFERROR(COUNTIFS({HRS},"<24")/COUNT({HRS}),0)','0.0%',bfont)
ws.cell(r+1,4,"Mediana sobre datos actuales. 'Completadas' = tasks creadas ese mes que ya están DONE.").font=Font(name="Arial",italic=True,size=8,color="808080")
for col,w in zip(['D','E','F','G','H','I','J'],[10,11,12,12,11,12,12]): ws.column_dimensions[col].width=w

# --- Tabla 2: distribución (done) ---
title('A15',"  DISTRIBUCIÓN POR TIEMPO DE RESOLUCIÓN (completadas)")
hdr(16,["Rango","# Tasks","% del total"])
buckets=[("Mismo día (< 24 h)",f'=COUNTIFS({HRS},"<24")'),("1 – 2 días",f'=COUNTIFS({HRS},">=24",{HRS},"<48")'),
 ("2 – 4 días",f'=COUNTIFS({HRS},">=48",{HRS},"<96")'),("4 – 7 días",f'=COUNTIFS({HRS},">=96",{HRS},"<168")'),
 ("Más de 7 días",f'=COUNTIFS({HRS},">=168")')]
r=17
for lbl,frm in buckets:
    box(r,1,lbl,'General',cfont,left); box(r,2,frm,'0'); box(r,3,f'=IFERROR(B{r}/COUNT({HRS}),0)','0.0%'); r+=1
box(r,1,"TOTAL",'General',bfont,left); box(r,2,'=SUM(B17:B21)','0',bfont); box(r,3,f'=IFERROR(B{r}/COUNT({HRS}),0)','0.0%',bfont)

# --- Tabla 3: por prioridad ---
title('D15',"  RESOLUTION TIME POR PRIORIDAD")
hdr(16,["Prioridad","# Tasks","% del total","Prom. días","Mediana días","% mismo día"],4)
prios=["URGENT","HIGH_PRIORITY","NORMAL"]
med_pri=done.groupby('Priority')['dias'].median().round(2).to_dict()
r=17
for p in prios:
    box(r,4,p,'General',cfont,left)
    box(r,5,f'=COUNTIFS({PRI},"{p}",{HRS},">=0")','0')
    box(r,6,f'=IFERROR(COUNTIFS({PRI},"{p}",{HRS},">=0")/COUNT({HRS}),0)','0.0%')
    box(r,7,f'=IFERROR(ROUND(AVERAGEIFS({DAYS},{PRI},"{p}"),1),0)','0.0')
    box(r,8, med_pri.get(p,0),'0.0')
    box(r,9,f'=IFERROR(COUNTIFS({PRI},"{p}",{HRS},"<24")/COUNTIFS({PRI},"{p}",{HRS},">=0"),0)','0.0%')
    r+=1
box(r,4,"TOTAL",'General',bfont,left); box(r,5,f'=COUNT({HRS})','0',bfont)
box(r,6,'100%','0%',bfont); box(r,7,f'=ROUND(AVERAGE({DAYS}),1)','0.0',bfont)
box(r,8,round(done['dias'].median(),1),'0.0',bfont); box(r,9,f'=IFERROR(COUNTIFS({HRS},"<24")/COUNT({HRS}),0)','0.0%',bfont)
ws.freeze_panes="A3"

# ===== Pendientes y En Proceso =====
op=df[df['Status'].isin(['TO_DO','WORKING'])].copy()
op['ab']=op['c'].apply(lambda x:max((TODAY.date()-x.date()).days,0))
op=op.sort_values('ab',ascending=False)
pz=wb.create_sheet("Pendientes y En Proceso"); pz.sheet_view.showGridLines=False
pz['A1']="TASKS ABIERTAS — Pendientes (TO_DO) y En Proceso (WORKING)"; pz['A1'].font=Font(name="Arial",bold=True,size=14,color=NAVY)
pz['A2']=f"{len(op)} tasks sin completar al {TODAY.date()}. Días abiertos = hoy − fecha de creación."
pz['A2'].font=Font(name="Arial",italic=True,size=9,color="808080")
cols=["Invoice Number","Status","Priority","Assigned by","Created At","Días abiertos"]
for c,h in enumerate(cols,1):
    cell=pz.cell(4,c,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
r=5
for _,row in op.iterrows():
    pz.cell(r,1,row['Invoice Number'] if pd.notna(row['Invoice Number']) else "")
    pz.cell(r,2,row['Status']); pz.cell(r,3,row['Priority'] if pd.notna(row['Priority']) else "")
    pz.cell(r,4,row['Assigned by'] if pd.notna(row['Assigned by']) else "")
    t=pz.cell(r,5,row['c'].tz_localize(None).to_pydatetime()); t.number_format='yyyy-mm-dd hh:mm'
    dd=pz.cell(r,6,int(row['ab'])); dd.font=bfont
    if row['ab']>=7: dd.fill=red
    elif row['ab']>=2: dd.fill=yel
    for c in range(1,7):
        cc=pz.cell(r,c); cc.border=border
        if c!=6: cc.font=cfont
        cc.alignment=left if c in (1,3,4) else center
    r+=1
for i,w in enumerate([16,12,15,16,18,14],1): pz.column_dimensions[get_column_letter(i)].width=w
pz.freeze_panes="A5"

# ===== Tasks lentas (>7 dias) =====
slow=done[done['dias']>=7].sort_values('dias',ascending=False)
sz=wb.create_sheet("Tasks lentas (>7 dias)"); sz.sheet_view.showGridLines=False
sz['A1']="TASKS COMPLETADAS QUE TARDARON MÁS DE 7 DÍAS"; sz['A1'].font=Font(name="Arial",bold=True,size=14,color=NAVY)
sz['A2']=f"{len(slow)} de {len(done)} tasks completadas ({round(100*len(slow)/len(done))}%) superaron los 7 días — sobre todo backlog histórico (dic-ene)."
sz['A2'].font=Font(name="Arial",italic=True,size=9,color="808080")
cols=["Invoice Number","Priority","Assigned by","Created At","Completed At","Días de resolución"]
for c,h in enumerate(cols,1):
    cell=sz.cell(4,c,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
r=5
for _,row in slow.iterrows():
    sz.cell(r,1,row['Invoice Number'] if pd.notna(row['Invoice Number']) else "")
    sz.cell(r,2,row['Priority'] if pd.notna(row['Priority']) else "")
    sz.cell(r,3,row['Assigned by'] if pd.notna(row['Assigned by']) else "")
    t=sz.cell(r,4,row['c'].tz_localize(None).to_pydatetime()); t.number_format='yyyy-mm-dd hh:mm'
    m=sz.cell(r,5,row['d'].tz_localize(None).to_pydatetime()); m.number_format='yyyy-mm-dd hh:mm'
    d=sz.cell(r,6); d.value=f'=(E{r}-D{r})'; d.number_format='0.0'; d.fill=red; d.font=bfont
    for c in range(1,7):
        cc=sz.cell(r,c); cc.border=border
        if c!=6: cc.font=cfont
        cc.alignment=left if c in (1,3) else center
    r+=1
sz.cell(r+1,5,"Promedio:").font=bfont
av=sz.cell(r+1,6,f'=ROUND(AVERAGE(F5:F{r-1}),1)'); av.font=bfont; av.number_format='0.0'
for i,w in enumerate([16,15,16,18,18,16],1): sz.column_dimensions[get_column_letter(i)].width=w
sz.freeze_panes="A5"

wb.calculation=CalcProperties(calcId=124519, fullCalcOnLoad=True)
wb.save(OUT)
print("OK ->",OUT,"| done:",len(done),"| abiertas:",len(op),"| lentas:",len(slow),"| meses:",len(meses))
