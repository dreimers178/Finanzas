#!/usr/bin/env python3
"""Genera el 'Reporte de Portal Uploads' a partir del export mensual.

Uso:  python3 generar_portal_uploads.py <archivo.xlsx|csv> [mes AAAA-MM] [salida.xlsx]
      - <archivo>: export con hoja 'Form Responses' (o CSV con esas columnas).
      - [mes]:     mes a destacar para distribución/portal/atípicos (default: último mes con datos).
      - [salida]:  ruta del xlsx a generar.

Métricas: resolution time = Message Timestamp − Timestamp.
Tablas: resolution por mes (todos los meses) · distribución (mes destacado) ·
        por portal normalizado (mes destacado) · atípicos >7 días (mes destacado).
"""
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

SRC = sys.argv[1] if len(sys.argv) > 1 else None
if SRC is None:
    sys.exit("Falta el archivo de entrada. Uso: generar_portal_uploads.py <archivo> [mes] [salida]")
TARGET = sys.argv[2] if len(sys.argv) > 2 else None
OUT = sys.argv[3] if len(sys.argv) > 3 else "/home/user/Finanzas/Reporte_Portal_Uploads.xlsx"

if SRC.lower().endswith('.csv'):
    resp = pd.read_csv(SRC)
    nom = pd.DataFrame()
else:
    resp = pd.read_excel(SRC, sheet_name='Form Responses')
    try: nom = pd.read_excel(SRC, sheet_name='Nomenclature')
    except Exception: nom = pd.DataFrame()

for col in ['Timestamp', 'Message Timestamp']:
    resp[col] = pd.to_datetime(resp[col], errors='coerce')
resp = resp[resp['Timestamp'].notna()].reset_index(drop=True)

def norm_portal(x):
    s = str(x).strip().lower()
    for k, v in [('ariba','Ariba'),('coupa','Coupa'),('chorus','Chorus Pro'),('msft','Microsoft'),
                 ('microsoft','Microsoft'),('taulia','Taulia'),('transcepta','Transcepta'),
                 ('google','Google'),('oracle','Oracle'),('tungsten','Tungsten')]:
        if k in s: return v
    if s.startswith('http'): return 'Portal por URL'
    if s in ('nan','','test'): return 'Sin dato'
    return str(x).strip().title()[:22]

calc = resp[resp['Message Timestamp'].notna()].copy()
calc['dias'] = (calc['Message Timestamp'] - calc['Timestamp']).dt.total_seconds() / 86400
calc['mes'] = calc['Timestamp'].dt.strftime('%Y-%m')
calc['portal'] = resp.loc[calc.index, 'Portal Name or Link'].apply(norm_portal)
if TARGET is None:
    TARGET = sorted(calc['mes'].unique())[-1]
tgt = calc[calc['mes'] == TARGET]
top_named = tgt['portal'].value_counts()
NAMED = [p for p, n in top_named.items() if n >= 2][:8]
otros = tgt[~tgt['portal'].isin(NAMED)]

NAVY="1F3864"
hfill=PatternFill("solid",fgColor=NAVY); sfill=PatternFill("solid",fgColor="2E5496")
hfont=Font(name="Arial",bold=True,color="FFFFFF",size=11); sfont=hfont
cfont=Font(name="Arial",size=10); bfont=Font(name="Arial",bold=True,size=10)
kfont=Font(name="Arial",bold=True,size=14,color=NAVY)
center=Alignment(horizontal="center",vertical="center"); left=Alignment(horizontal="left",vertical="center")
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
red=PatternFill("solid",fgColor="F4CCCC")

wb=Workbook()

def dump(ws, dfin, fmt_ts=None):
    for c,h in enumerate(dfin.columns,1):
        cell=ws.cell(1,c,str(h)); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
    for r,(_,row) in enumerate(dfin.iterrows(),2):
        for c,h in enumerate(dfin.columns,1):
            v=row[h]
            if pd.isna(v): v=None
            elif isinstance(v,pd.Timestamp): v=v.to_pydatetime()
            elif isinstance(v,str): v=v.replace('\n',' / ')
            cell=ws.cell(r,c,v); cell.font=cfont
            if isinstance(row[h],pd.Timestamp): cell.number_format='yyyy-mm-dd hh:mm'
    for i in range(1,len(dfin.columns)+1): ws.column_dimensions[get_column_letter(i)].width=18
    ws.freeze_panes="A2"

ws1=wb.active; ws1.title="Form Responses"; dump(ws1, resp)
if not nom.empty:
    ws2=wb.create_sheet("Nomenclature")
    nom2=nom.drop(columns=[c for c in nom.columns if str(c).startswith('Unnamed')],errors='ignore')
    dump(ws2, nom2)

# Datos (calculo)
dz=wb.create_sheet("Datos (calculo)")
dcols=["Invoice Number","Submitted By","Collector assigned","Submit Timestamp","Resolved Timestamp",
       "Resolution (hrs)","Resolution (dias)","Mes","Portal (norm.)"]
for c,h in enumerate(dcols,1):
    cell=dz.cell(1,c,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
r=2
for _,row in resp.iterrows():
    inv=str(row['Invoice Number']).replace('\n',' / ') if pd.notna(row['Invoice Number']) else ""
    dz.cell(r,1,inv); dz.cell(r,2,row.get('Submitted By') if pd.notna(row.get('Submitted By')) else "")
    dz.cell(r,3,row.get('Collector assigned') if pd.notna(row.get('Collector assigned')) else "")
    ts=dz.cell(r,4,row['Timestamp'].to_pydatetime()); ts.number_format='yyyy-mm-dd hh:mm'
    if pd.notna(row['Message Timestamp']):
        mt=dz.cell(r,5,row['Message Timestamp'].to_pydatetime()); mt.number_format='yyyy-mm-dd hh:mm'
        h=(row['Message Timestamp']-row['Timestamp']).total_seconds()/3600
        dz.cell(r,6,round(h,1)); dz.cell(r,7,round(h/24,1))
    dz.cell(r,6).number_format='0.0'; dz.cell(r,7).number_format='0.0'
    dz.cell(r,8,row['Timestamp'].strftime('%Y-%m')); dz.cell(r,9,norm_portal(row['Portal Name or Link']))
    for c in range(1,10):
        cc=dz.cell(r,c); cc.font=cfont; cc.border=border; cc.alignment=left if c in (1,2,3,9) else center
    r+=1
LAST=r-1
for i,w in enumerate([26,22,16,20,20,15,15,10,18],1): dz.column_dimensions[get_column_letter(i)].width=w
dz.freeze_panes="A2"
D="'Datos (calculo)'"; HRS=f"{D}!$F$2:$F${LAST}"; DAYS=f"{D}!$G$2:$G${LAST}"
MES=f"{D}!$H$2:$H${LAST}"; PRT=f"{D}!$I$2:$I${LAST}"

ws=wb.create_sheet(f"Reporte {TARGET}", 0)
ws.sheet_view.showGridLines=False
def title(cell,txt): ws[cell]=txt; ws[cell].font=sfont; ws[cell].fill=sfill; ws[cell].alignment=left
def hdr(row,cols,sc=1):
    for i,h in enumerate(cols):
        cell=ws.cell(row,sc+i,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
def box(row,col,val,fmt='General',font=cfont,al=center):
    cell=ws.cell(row,col,val); cell.font=font; cell.alignment=al; cell.number_format=fmt; cell.border=border; return cell

ws['A1']=f"REPORTE DE PORTAL UPLOADS — {TARGET}"; ws['A1'].font=Font(name="Arial",bold=True,size=16,color=NAVY)
ws['A2']=f"Resolution Time = Resolved − Submit. Distribución/portal/atípicos filtrados a {TARGET}; la tabla mensual muestra todos los meses."
ws['A2'].font=Font(name="Arial",italic=True,size=9,color="808080")

title('A4',"  RESOLUTION TIME POR MES")
hdr(5,["Mes","# Uploads","Resueltos","Prom. horas","Prom. días","% mismo día"])
meses=sorted(calc['mes'].unique()); r=6
for m in meses:
    box(r,1,m,'General',bfont,left); box(r,2,f'=COUNTIF({MES},"{m}")','0')
    box(r,3,f'=COUNTIFS({MES},"{m}",{HRS},">=0")','0')
    box(r,4,f'=IFERROR(ROUND(AVERAGEIF({MES},"{m}",{HRS}),1),0)','0.0')
    box(r,5,f'=IFERROR(ROUND(AVERAGEIF({MES},"{m}",{DAYS}),1),0)','0.0')
    box(r,6,f'=IFERROR(COUNTIFS({MES},"{m}",{HRS},"<24")/COUNTIFS({MES},"{m}",{HRS},">=0"),0)','0.0%'); r+=1
box(r,1,"TOTAL",'General',bfont,left); box(r,2,f'=SUM(B6:B{r-1})','0',bfont); box(r,3,f'=SUM(C6:C{r-1})','0',bfont)
box(r,4,f'=ROUND(AVERAGE({HRS}),1)','0.0',bfont); box(r,5,f'=ROUND(AVERAGE({DAYS}),1)','0.0',bfont)
box(r,6,f'=IFERROR(COUNTIFS({HRS},"<24")/COUNT({HRS}),0)','0.0%',bfont)
for col,w in zip(['A','B','C','D','E','F'],[12,12,12,13,13,13]): ws.column_dimensions[col].width=w

d0=r+3
title(f'A{d0}',f"  DISTRIBUCIÓN POR TIEMPO DE RESOLUCIÓN — {TARGET}")
hdr(d0+1,["Rango","# Uploads","% del total"])
buckets=[("Mismo día (< 24 h)",f'"<24"'),("1 – 2 días",f'">=24",{HRS},"<48"'),("2 – 4 días",f'">=48",{HRS},"<96"'),
 ("4 – 7 días",f'">=96",{HRS},"<168"'),("Más de 7 días",f'">=168"')]
rr=d0+2; first=rr
for lbl,cond in buckets:
    box(rr,1,lbl,'General',cfont,left); box(rr,2,f'=COUNTIFS({MES},"{TARGET}",{HRS},{cond})','0')
    box(rr,3,f'=IFERROR(B{rr}/COUNTIFS({MES},"{TARGET}",{HRS},">=0"),0)','0.0%'); rr+=1
box(rr,1,"TOTAL",'General',bfont,left); box(rr,2,f'=SUM(B{first}:B{rr-1})','0',bfont)
box(rr,3,f'=IFERROR(B{rr}/COUNTIFS({MES},"{TARGET}",{HRS},">=0"),0)','0.0%',bfont)

p0=rr+3
title(f'A{p0}',f"  RESOLUTION TIME POR PORTAL — {TARGET}")
hdr(p0+1,["Portal","# Uploads","% del total","Prom. días","% mismo día"])
TOT=f'COUNTIFS({MES},"{TARGET}",{HRS},">=0")'; rr=p0+2
for p in NAMED:
    box(rr,1,p,'General',cfont,left); box(rr,2,f'=COUNTIFS({PRT},"{p}",{MES},"{TARGET}")','0')
    box(rr,3,f'=IFERROR(COUNTIFS({PRT},"{p}",{MES},"{TARGET}")/{TOT},0)','0.0%')
    box(rr,4,f'=IFERROR(ROUND(AVERAGEIFS({DAYS},{PRT},"{p}",{MES},"{TARGET}"),1),0)','0.0')
    box(rr,5,f'=IFERROR(COUNTIFS({PRT},"{p}",{MES},"{TARGET}",{HRS},"<24")/COUNTIFS({PRT},"{p}",{MES},"{TARGET}"),0)','0.0%'); rr+=1
box(rr,1,"Otros",'General',cfont,left); box(rr,2,len(otros),'0'); box(rr,3,f'=IFERROR(B{rr}/{TOT},0)','0.0%')
box(rr,4,round(otros['dias'].mean(),1) if len(otros) else 0,'0.0'); box(rr,5,round((otros['dias']<1).mean(),3) if len(otros) else 0,'0.0%'); rr+=1
box(rr,1,"TOTAL",'General',bfont,left); box(rr,2,f'={TOT}','0',bfont); box(rr,3,f'=IFERROR(B{rr}/{TOT},0)','0.0%',bfont)
box(rr,4,f'=ROUND(AVERAGEIF({MES},"{TARGET}",{DAYS}),1)','0.0',bfont); box(rr,5,f'=IFERROR(COUNTIFS({MES},"{TARGET}",{HRS},"<24")/{TOT},0)','0.0%',bfont)
ws.freeze_panes="A3"

out=tgt[tgt['dias']>=7].sort_values('dias',ascending=False)
az=wb.create_sheet(f"Atipicos {TARGET} (>7d)"); az.sheet_view.showGridLines=False
az['A1']=f"CASOS ATÍPICOS DE {TARGET} — Resolución mayor a 7 días"; az['A1'].font=Font(name="Arial",bold=True,size=14,color=NAVY)
az['A2']=f"{len(out)} uploads de {TARGET} superaron los 7 días."; az['A2'].font=Font(name="Arial",italic=True,size=9,color="808080")
cols=["Invoice Number","Submitted By","Collector assigned","Portal","Submit Timestamp","Resolved Timestamp","Días de resolución"]
for c,h in enumerate(cols,1):
    cell=az.cell(4,c,h); cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
r=5
for _,row in out.iterrows():
    az.cell(r,1,str(row['Invoice Number']).replace('\n',' / ')); az.cell(r,2,row.get('Submitted By'))
    az.cell(r,3,row.get('Collector assigned')); az.cell(r,4,row['portal'])
    t=az.cell(r,5,row['Timestamp'].to_pydatetime()); t.number_format='yyyy-mm-dd hh:mm'
    m=az.cell(r,6,row['Message Timestamp'].to_pydatetime()); m.number_format='yyyy-mm-dd hh:mm'
    d=az.cell(r,7); d.value=f'=(F{r}-E{r})'; d.number_format='0.0'; d.fill=red; d.font=bfont
    for c in range(1,8):
        cc=az.cell(r,c); cc.border=border
        if c!=7: cc.font=cfont
        cc.alignment=left if c in (1,2,3,4) else center
    r+=1
if len(out):
    az.cell(r+1,6,"Promedio atípicos:").font=bfont
    av=az.cell(r+1,7,f'=ROUND(AVERAGE(G5:G{r-1}),1)'); av.font=bfont; av.number_format='0.0'
for i,w in enumerate([24,22,16,16,20,20,16],1): az.column_dimensions[get_column_letter(i)].width=w
az.freeze_panes="A5"

wb.calculation=CalcProperties(calcId=124519, fullCalcOnLoad=True)
wb.save(OUT)
print("OK ->",OUT,"| mes destacado:",TARGET,"| meses:",meses,"| atipicos:",len(out))
